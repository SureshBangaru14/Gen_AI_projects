import os
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

from generate_text import GenerateText


# ----------------------------------
# Load environment
# ----------------------------------
load_dotenv("/home/suresh/Gen_AI_Practice/.env")

base_path = os.getenv("RESUME_RAW_DATA")

if not base_path:
    raise ValueError("RESUME_RAW_DATA is not set in the .env file.")

base_path = Path(base_path)

if not base_path.exists():
    raise FileNotFoundError(f"Path does not exist: {base_path}")


# ----------------------------------
# Output paths
# ----------------------------------
output_file = Path(
    "/home/suresh/generative-ai/Generative-AI/projects/resume_ats/data/resume_dataset.xlsx"
)

txt_output_dir = Path(
    "/home/suresh/generative-ai/Generative-AI/projects/resume_ats/data/resume_text"
)

output_file.parent.mkdir(parents=True, exist_ok=True)
txt_output_dir.mkdir(parents=True, exist_ok=True)

sheet_name = "Resume Data"


# ----------------------------------
# Tracking
# ----------------------------------
total = 0
success = 0
skipped = 0
failed = 0

skipped_files = []
failed_files = []
processed_files_list = []


# ----------------------------------
# Create Excel if not exists
# ----------------------------------
if not output_file.exists():
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(["Category", "File Name", "Extracted Text"])
    wb.save(output_file)
    wb.close()


# ----------------------------------
# Load workbook
# ----------------------------------
wb = load_workbook(output_file)

if sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
else:
    ws = wb.create_sheet(sheet_name)
    ws.append(["Category", "File Name", "Extracted Text"])


# ----------------------------------
# Avoid duplicates
# ----------------------------------
processed_files = {
    (row[0], row[1])
    for row in ws.iter_rows(min_row=2, values_only=True)
}


# ----------------------------------
# Process PDFs
# ----------------------------------
for category_dir in sorted(base_path.iterdir()):

    if not category_dir.is_dir():
        continue

    category = category_dir.name
    safe_category = category.replace(" ", "_").strip()

    for pdf_file in sorted(category_dir.glob("*.pdf")):

        total += 1
        key = (category, pdf_file.name)

        # Skip duplicates
        if key in processed_files:
            skipped += 1
            skipped_files.append(str(pdf_file))
            print(f"Skipped: {pdf_file.name}")
            continue

        try:
            print(f"Processing: {pdf_file.name}")

            # --------------------------
            # Extract text
            # --------------------------
            text = GenerateText.extract_with_ocr(str(pdf_file))

            if not text:
                text = ""

            text = str(text).strip()

            # --------------------------
            # Handle Excel limit
            # --------------------------
            MAX_EXCEL_LIMIT = 32767

            if len(text) > MAX_EXCEL_LIMIT:
                text = text[1000:]  # remove noise only if large

            excel_text = text[:MAX_EXCEL_LIMIT]

            # --------------------------
            # Save FULL TEXT to TXT
            # --------------------------
            txt_file = txt_output_dir / f"{safe_category}_{pdf_file.stem}.txt"

            with open(txt_file, "w", encoding="utf-8") as f:
                f.write(text)

            # --------------------------
            # Save to Excel
            # --------------------------
            ws.append([
                category,
                pdf_file.name,
                excel_text
            ])

            processed_files.add(key)
            success += 1
            processed_files_list.append(str(pdf_file))

            wb.save(output_file)

            print(f"Saved Excel: {pdf_file.name}")
            print(f"Saved TXT  : {txt_file.name}")

        except Exception as e:
            failed += 1
            failed_files.append(str(pdf_file))
            print(f"Error: {pdf_file.name} -> {type(e).__name__}: {e}")


# ----------------------------------
# Final save
# ----------------------------------
wb.save(output_file)
wb.close()


# ----------------------------------
# Summary
# ----------------------------------
print("\n===================================")
print("Resume extraction completed")
print("===================================")

print(f"Total PDFs     : {total}")
print(f"Processed      : {success}")
print(f"Skipped        : {skipped}")
print(f"Failed         : {failed}")

print(f"\nExcel Location : {output_file}")
print(f"TXT Folder     : {txt_output_dir}")

print("\n---------- PROCESSED FILES ----------")
for f in processed_files_list:
    print(f"✔ {f}")

print("\n---------- SKIPPED FILES ----------")
for f in skipped_files:
    print(f"⚠ {f}")

print("\n---------- FAILED FILES ----------")
for f in failed_files:
    print(f"❌ {f}")

print("===================================")